from openpyxl import*
from openpyxl.utils import*
import datetime as dt
colone = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA",
          "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ"]
nom=input("FNC?\n")

# chemin=""
# if nom[7:9]=="01":
#     chemin="Y:/05-Quality/02-Non Conformités/03-Enregistrements_Non Conformités/01-En cours de traitement/01-Provenance fournisseur/"
# if nom[7:9]=="02":
#     chemin="Y:/05-Quality/02-Non Conformités/03-Enregistrements_Non Conformités/01-En cours de traitement/02-Detecté en Interne/"
# if nom[7:9]=="03":
#     chemin="Y:/05-Quality/02-Non Conformités/03-Enregistrements_Non Conformités/01-En cours de traitement/03-Détecté par le client/"
feuille = load_workbook(nom+'.xlsx', data_only=True)
monitoring = load_workbook('FR 1005-Monitoring NC & Action Plan-V001.xlsx')
checklist = load_workbook('checklist.xlsm', keep_vba=True)
mo=monitoring["Suivi_FNC"]
su=monitoring["Plan_Actions"]
nc = feuille["Fiche de Non-Conformité"]
ch=checklist["INT_DATA"]
ch_ligne=mo_ligne=su_ligne=1
all_fnc=[]
all_pl=[]
while ch[colone[1]+str(ch_ligne)].data_type == "s":
    ch_ligne += 1
while mo[colone[0]+str(mo_ligne)].data_type == "s":
    mo_ligne += 1
while su[colone[0]+str(su_ligne)].data_type == "s":
    su_ligne += 1
for i in range(3,mo_ligne):
    all_fnc.append(mo[colone[0]+str(i)].value)
for i in range (2,su_ligne):
    all_pl.append([su[colone[0]+str(i)].value,su[colone[1]+str(i)].value])

données=[]
données.append(nc["l7"].value)
données.append (dt.datetime.strftime((nc["AJ7"].value), "%d/%m/%Y"))
données.append (str((nc["S10"].value)).upper()+' '+ (nc["F10"].value))
détection=""
if(nc["AT7"].value)==1 :
    détection="Fournisseur"
if(nc["AT7"].value)==2 :
    détection="Interne"
if(nc["AT7"].value)==3 :
    détection="Client"
données.append(détection)
données.append("")
données.append(nc["AN16"].value) 
données.append(nc["C13"].value)
données.append(nc["AH13"].value)
#####section 3 
données.append(nc["j32"].value) 
données.append(nc["p32"].value) 
données.append(nc["v32"].value) 
données.append(nc["ak32"].value)
#####section 4
solution=""
if(nc["d38"].value)!="":
    if solution=="":
        solution+="MFT "
    else:
        solution+="+ MFT "
else:
    solution+=""

if(nc["k40"].value)!="":
    if solution=="":
        solution+="5 Why "
    else:
        solution+="+ 5 Why "
else:
    solution+=""
if(nc["m50"].value)!="":
    if solution=="":
        solution+="Ishikawa "
    else:
        solution+="+ Ishikawa "
else:
    solution+=""
données.append(solution)
données.append(nc["d38"].value+nc["k40"].value+nc["m50"].value)
données.append(nc["au9"].value)
#### plan d'action
plan_action=[]
num_action=[21,22,23,62,63,64,65,66]
for i in range (len(num_action)):
    if nc["B"+str(num_action[i])].data_type!="n":
        plan_action.append([nc["B"+str(num_action[i])].value,nom[:12],"",nc["G"+str(num_action[i])].value,"",nc["AF"+str(num_action[i])].value,dt.datetime.strftime((nc["AP"+str(num_action[i])].value), "%d/%m/%Y")])
for i in range (len(plan_action)):
    for j in range (len(plan_action[i])):
        if plan_action[i][:2] not in all_pl:
            su[colone[j]+str(su_ligne)]=plan_action[i][j]
    su_ligne+=1
#####insertion données
for i in range (len(all_fnc)):
    if nom[:12] == all_fnc[i]:
        mo_ligne=i+3
for i in range(len(données)):
        mo[colone[i]+str(mo_ligne)] = données[i]
if nom[:12] not in all_fnc:
    ch[colone[1]+str(ch_ligne)] = données[0]

monitoring.save('FR 1005-Monitoring NC & Action Plan-V001.xlsx')
checklist.save('checklist.xlsm')