from openpyxl import*
from openpyxl.utils import*
import pandas as pd
import datetime as dt

from requests import delete
df = pd.read_excel('fac.xls', header=None)
df.to_excel('fac.xlsx', index=False, header=False)
read_file = pd.read_csv(
    r'C:/Users/Etudiant/Desktop/stage_excel/ExpDevis.csv', delimiter=";", encoding='latin1')
read_file.to_excel(
    r'C:/Users/Etudiant/Desktop/stage_excel/ExpDevis.xlsx', index=None, header=True)
export = load_workbook('ExpDevis.xlsx')
facture = load_workbook('fac.xlsx')
data = load_workbook('HYFR_DC-FC_2022 JZ.xlsx')
ex = export.active
fa = facture.active
da = data["Data"]
colone = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA",
          "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ"]
ligne_ex = 2
while ex[colone[0]+str(ligne_ex)].data_type == "s":
    ligne_ex += 1
all_devis = []
tmp_devis = ()

################################ données devis ################################
for i in range(2, ligne_ex):
    ligne_data = 3
    while da[colone[6]+str(ligne_data)].data_type == "s":
        ligne_data += 1
    all_index = []
    ind = 0
    for ind in range(ligne_data):
        if type(da[colone[4]+str(ind+1)].value) == int:
            all_index.append(da[colone[4]+str(ind+1)].value)
    for devis in range(3, (ligne_data)):
        tmp_devis = (dt.datetime.strftime((da[colone[3]+str(devis)].value), "%d/%m/%Y"), (da[colone[4] + str(devis)].value), (da[colone[6] + str(devis)].value), (da[colone[8] + str(devis)].value),
                     (da[colone[9]+str(devis)].value), (da[colone[11] + str(devis)].value), (da[colone[17]+str(devis)].value), ((da[colone[18] + str(devis)].value)),
                     (da[colone[26]+str(devis)].value))
        all_devis.append(tmp_devis)
    date_devis = dt.datetime.strptime((ex[colone[0]+str(i)].value), "%d/%m/%Y")
    bu_key = f"=IFERROR(INDEX(Tabelle2[BU],MATCH([[#This Row],[Categorie]],CAT,0)),\"\")"
    cw = f"=IF({'D' + str(ligne_data)}<>\"\",\"S\"&TEXT(WEEKNUM({'D' + str(ligne_data)},21),\"00\"),\"\")"
    N_devis = (ex[colone[1]+str(i)].value)
    Client = (ex[colone[2]+str(i)].value)
    ASM = (ex[colone[3]+str(i)].value)
    Montant_devis = float(ex[colone[4]+str(i)].value)
    Date = date_devis
    Status = None
    if (ex[colone[5]+str(i)].data_type) == "s":
        Status = 'Accepté'
    if Montant_devis == 0:
        Status = 'Opportunité'
    indice_devis = all_index.count(N_devis)
    index = colone[indice_devis]
    Client_end = None
    Catégorie = None
    données_devis = [Date, bu_key, cw, date_devis, N_devis,
                     index, Client, Client_end, Montant_devis, ASM, Catégorie, Status]

    ################################ Split ################################
    split = ""

    ################################ données commande ################################
    if (ex[colone[5]+str(i)].data_type) == "s":
        cw_order = f"=IF({'Q' + str(ligne_data)}<>\"\",\"S\"&TEXT(WEEKNUM({'Q' + str(ligne_data)},21),\"00\"),\"\")"
        OI_Year = f"=IF(Data!${'Q'+ str(ligne_data)}=\"\",\"\",YEAR(Data!${'Q'+ str(ligne_data)}))"
        OI_Month = f"=IF(Data!${'Q'+ str(ligne_data)}=\"\",\"\",MONTH(Data!${'Q'+ str(ligne_data)}))"
        date_commande = (ex[colone[5]+str(i)].value)
        n_commande = (ex[colone[6]+str(i)].value)
        montant_commande = float(ex[colone[7]+str(i)].value)
        montant_restant = None
        délai = None
    else:
        cw_order = None
        OI_Year = None
        OI_Month = None
        date_commande = None
        n_commande = None
        montant_commande = None
        montant_restant = None
        délai = None
    données_commande = [split, cw_order, OI_Year, OI_Month,
                        date_commande, n_commande, montant_commande, montant_restant, délai]
    ################################ Split2 ################################
    split2 = ""

    ################################ données facture ################################
    if (ex[colone[9]+str(i)].data_type) == "s":
        date_réelle = " "
        cw_rev = f"=IF({'AC'+ str(ligne_data)}<>\"\",\"S\"&TEXT(WEEKNUM({'AC'+ str(ligne_data)},21),\"00\"),\"\")"
        revenue_Year = f"=IF(Data!${'AC'+ str(ligne_data)}=\"\",\"\",YEAR(Data!${'AC'+ str(ligne_data)}))"
        revenue_Month = f"=IF(Data!${'AC'+ str(ligne_data)}=\"\",\"\",MONTH(Data!${'AC'+ str(ligne_data)}))"
        dateliv_réelle = (ex[colone[8]+str(i)].value)
        date_facture = (ex[colone[9]+str(i)].value)
        n_facture = (Client, date_facture, montant_commande)
        ligne_fac = 8
        while fa[colone[4]+str(ligne_fac)].data_type == "d":
            ligne_fac += 1
        num_fac = {}
        i = 8
        # récup des num de facture pour les clés et les dates,montants et cliens
        for i in range(i, ligne_fac):
            num_fac[fa[colone[0]+str(i)].value] = (fa[colone[1]+str(i)].value, dt.datetime.strftime(
                (fa[colone[4]+str(i)].value), "%d/%m/%Y"), fa[colone[11]+str(i)].value,)
            i += 1
        # recherche du num de facture dans les clés uniques
        if n_facture in list(num_fac.values()):
            n_facture = (list(num_fac.keys())[
                         list(num_fac.values()).index(n_facture)])
        else:
            n_facture = None
    else:
        date_réelle = None
        cw_rev = None
        revenue_Year = None
        revenue_Month = None
        dateliv_réelle = None
        date_facture = None
        n_facture = None
    données_facture = [split2, date_réelle, cw_rev, revenue_Year,
                       revenue_Month, dateliv_réelle, n_facture, date_facture]
    données = données_devis+données_commande+données_facture
    ##############################  vérif devis  ################################
    verif_devis = (dt.datetime.strftime(date_devis, "%d/%m/%Y"), N_devis, Client, float(Montant_devis), ASM, Status,
                   n_commande, montant_commande, dateliv_réelle)
    ################################ insersion excel ################################
    if verif_devis not in all_devis:
        for i in range(len(all_devis)) :
            if verif_devis[:5]==all_devis[i][:5] and (all_devis[i][8]==None or all_devis[i][6]==None):
                da.delete_rows(i+3)
                données[5]=colone[indice_devis-1]
                ligne_data-=1 
                print("doublons")     
        print(verif_devis)
        print(all_devis[len(all_devis)-1])
        # print("\n\n\n",all_devis)
        for i in range(len(données)):
            da[colone[i]+str(ligne_data)] = données[i]
        ligne_data += 1
    else:
        print("les données sont identiques")
data.save('HYFR_DC-FC_2022 JZ.xlsx')
